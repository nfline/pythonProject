import requests
import openpyxl
from openpyxl.styles import PatternFill
import base64
import concurrent.futures
import logging
from datetime import datetime, timedelta
import time
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from threading import Semaphore, Lock
import urllib3
import psutil
import warnings
import ssl
import os
import pandas as pd
from collections import defaultdict

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings()

try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context

# Setup and API credentials
HOST = (".api.cloud.extrahop.com")
ID = ""
SECRET = ""
EXCEL_FILE = "device.xlsx"
MAX_WORKERS = 20  # Increased from 10 to 20
BATCH_SIZE_INITIAL = 200  # Increased initial batch size
MAX_BATCH_SIZE = 500  # Increased maximum batch size
CONNECTION_TIMEOUT = 30  # Connection timeout in seconds
SESSION_POOL_SIZE = 100  # Connection pool size

# Token caching
class TokenCache:
    def __init__(self):
        self.token = None
        self.expiry = None
        self.lock = Lock()

    def is_valid(self):
        with self.lock:
            return self.token and self.expiry and datetime.now() < self.expiry

    def set_token(self, token, expires_in=3600):
        with self.lock:
            self.token = token
            self.expiry = datetime.now() + timedelta(seconds=expires_in)

token_cache = TokenCache()

# Session pool
session_pool = []
session_pool_lock = Lock()

# Setup logging
def setup_logging():
    log_filename = f"tagging_operations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler()
        ]
    )
    return log_filename

# Global tracking variables
class TaggingStats:
    def __init__(self):
        self.successful_name_tags = set()
        self.failed_name_tags = set()
        self.successful_ip_tags = set()
        self.failed_ip_tags = set()
        self.created_tags = set()
        self.existing_tags = set()
        self.processed_count = 0
        self.start_time = None
        self.tag_cache = {}  # Cache for tag IDs
        self.lock = Lock()

    def increment_processed(self):
        with self.lock:
            self.processed_count += 1
            if self.processed_count % 1000 == 0:
                elapsed = datetime.now() - self.start_time
                rate = self.processed_count / elapsed.total_seconds() if elapsed.total_seconds() > 0 else 0
                logging.info(f"Processed {self.processed_count} items. Rate: {rate:.2f} items/sec")

stats = TaggingStats()

def get_session():
    """Get a session from the pool or create a new one"""
    with session_pool_lock:
        if session_pool:
            return session_pool.pop()
        else:
            return create_session()

def return_session(session):
    """Return a session to the pool"""
    with session_pool_lock:
        if len(session_pool) < SESSION_POOL_SIZE:
            session_pool.append(session)

def create_session():
    """Create a requests session with retry strategy"""
    session = requests.Session()
    retry_strategy = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    adapter = HTTPAdapter(
        max_retries=retry_strategy,
        pool_connections=10,
        pool_maxsize=100
    )
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.verify = False  # Disable SSL verification
    return session

def get_token():
    """Generate and retrieve a temporary API access token with caching and retry logic."""
    max_retries = 3
    retry_delay = 2  # seconds

    for attempt in range(max_retries):
        if token_cache.is_valid():
            return token_cache.token

        try:
            auth = f"{ID}:{SECRET}".encode('utf-8')
            headers = {
                "Authorization": f"Basic {base64.b64encode(auth).decode('utf-8')}",
                "Content-Type": "application/x-www-form-urlencoded"
            }
            url = f"https://{HOST}/oauth2/token"
            session = get_session()
            try:
                response = session.post(url, headers=headers, data="grant_type=client_credentials", timeout=CONNECTION_TIMEOUT)
                
                if response.status_code == 200:
                    try:
                        token_data = response.json()
                        token = token_data.get('access_token')
                        expires_in = token_data.get('expires_in', 3600)
                        
                        if not token:
                            logging.error("Token not found in response")
                            time.sleep(retry_delay)
                            continue
                            
                        # Refresh token 5 minutes before expiry
                        token_cache.set_token(token, expires_in - 300)
                        return token
                    except requests.exceptions.JSONDecodeError:
                        logging.error("Error decoding token response JSON")
                else:
                    logging.error(f"Failed to get token: {response.status_code}, Response: {response.text}")
                
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
            finally:
                return_session(session)
                    
        except (requests.exceptions.RequestException, ValueError) as e:
            logging.error(f"Error during token retrieval: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)

    return None

def get_auth_header():
    """Retrieve the authorization header with the token."""
    token = get_token()
    if not token:
        logging.error("Authentication failed: No token retrieved.")
        return None
    return f"Bearer {token}"

def get_tag_id(tag_name, headers):
    """Retrieve the ID of a tag based on its name with caching."""
    # Check the cache first
    if tag_name in stats.tag_cache:
        return stats.tag_cache[tag_name]
    
    session = get_session()
    try:
        url = f"https://{HOST}/api/v1/tags"
        response = session.get(url, headers=headers)
        if response.status_code != 200:
            logging.error(f"Failed to fetch tags: {response.status_code}, Response: {response.text}")
            return None
        
        try:
            tags = response.json()
            tag_id = next((tag['id'] for tag in tags if tag['name'] == tag_name), None)
            if tag_id:
                logging.info(f"Found existing tag: '{tag_name}' with ID: {tag_id}")
                stats.existing_tags.add(tag_name)
                # Cache the tag ID
                stats.tag_cache[tag_name] = tag_id
            else:
                logging.info(f"Tag '{tag_name}' not found in API response.")
            return tag_id
        except requests.exceptions.JSONDecodeError:
            logging.error("Error decoding tags response JSON.")
            return None
    finally:
        return_session(session)

def create_tag(tag_name, headers):
    """Create a new tag if it does not exist."""
    session = get_session()
    try:
        url = f"https://{HOST}/api/v1/tags"
        data = {"name": tag_name}
        response = session.post(url, headers=headers, json=data)
        if response.status_code != 201:
            logging.error(f"Failed to create tag '{tag_name}': {response.status_code}, Response: {response.text}")
            return None

        try:
            tag_id = response.json().get('id')
        except requests.exceptions.JSONDecodeError:
            logging.warning("No JSON in response, checking headers for tag ID...")
            tag_id = None

        if not tag_id and 'Location' in response.headers:
            tag_id = response.headers['Location'].split('/')[-1]
            logging.info(f"Extracted tag ID from Location header: {tag_id}")

        if not tag_id:
            logging.error(f"Created tag '{tag_name}', but could not retrieve its ID!")
            return None

        logging.info(f"Successfully created new tag '{tag_name}' with ID: {tag_id}")
        stats.created_tags.add(tag_name)
        # Cache the tag ID
        stats.tag_cache[tag_name] = tag_id
        return tag_id
    finally:
        return_session(session)

def check_memory_usage():
    """Monitor system memory usage"""
    process = psutil.Process()
    memory_info = process.memory_info()
    memory_percent = process.memory_percent()
    if memory_percent > 80:  # If memory usage exceeds 80%
        logging.warning(f"High memory usage detected: {memory_percent:.2f}% ({memory_info.rss / 1024 / 1024:.2f} MB)")
        return True
    return False

def refresh_token_if_needed(headers):
    """Check and refresh token if expired or about to expire"""
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            if not token_cache.is_valid():
                new_token = get_token()
                if new_token:
                    headers["Authorization"] = f"Bearer {new_token}"
                    logging.info("Token refreshed successfully")
                    return True
                else:
                    if attempt < max_retries - 1:
                        logging.warning(f"Token refresh attempt {attempt + 1} failed, retrying...")
                        time.sleep(retry_delay)
                        continue
                    else:
                        logging.error("All token refresh attempts failed")
                        return False
            return True
        except Exception as e:
            logging.error(f"Error during token refresh: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                continue
            return False
    
    return False

def batch_search_and_tag_devices(values, field, tag_id, headers, cells):
    """Batch search and tag devices based on field and values."""
    session = get_session()
    try:
        url_search = f"https://{HOST}/api/v1/devices/search"
        
        # Track success/failure for each device
        success_values = []
        failed_values = []
        success_cells = []
        failed_cells = []
        
        # Check memory usage
        if check_memory_usage():
            time.sleep(2)  # Allow system time for memory cleanup
        
        # Refresh token if needed
        if not refresh_token_if_needed(headers):
            logging.error("Token refresh failed, marking batch as failed")
            for cell in cells:
                cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
            return False
        
        # Search devices in batch
        all_device_ids = []
        for value, cell in zip(values, cells):
            try:
                data = {"filter": {"field": field, "operand": value, "operator": "="}}
                response = session.post(url_search, headers=headers, json=data, timeout=CONNECTION_TIMEOUT)
                
                if response.status_code == 200:
                    try:
                        devices = response.json()
                        stats.increment_processed()
                        if devices:
                            all_device_ids.extend([device['id'] for device in devices])
                            success_values.append(value)
                            success_cells.append(cell)
                        else:
                            failed_values.append(value)
                            failed_cells.append(cell)
                    except requests.exceptions.JSONDecodeError:
                        logging.error(f"Error decoding device search response JSON for {field} {value}")
                        failed_values.append(value)
                        failed_cells.append(cell)
                else:
                    failed_values.append(value)
                    failed_cells.append(cell)
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                logging.error(f"Request failed for {field} {value}: {str(e)}")
                failed_values.append(value)
                failed_cells.append(cell)
        
        # Mark failed cells
        for cell in failed_cells:
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
        
        if not all_device_ids:
            return False

        # Split large device ID lists into smaller chunks to avoid API limits
        device_id_chunks = [all_device_ids[i:i+500] for i in range(0, len(all_device_ids), 500)]
        
        # Refresh token before tag assignment
        if not refresh_token_if_needed(headers):
            logging.error("Token refresh failed before tag assignment")
            return False

        # Assign tags in batches
        success_count = 0
        for device_chunk in device_id_chunks:
            try:
                url_assign = f"https://{HOST}/api/v1/tags/{tag_id}/devices"
                data_assign = {"assign": device_chunk}
                response_assign = session.post(url_assign, headers=headers, json=data_assign, timeout=CONNECTION_TIMEOUT)
                
                if response_assign.status_code == 204:
                    success_count += len(device_chunk)
                else:
                    logging.error(f"Failed to assign tag to chunk of {len(device_chunk)} devices: {response_assign.status_code}")
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                logging.error(f"Tag assignment request failed for chunk: {str(e)}")
        
        # Only consider the batch successful if at least some devices were tagged
        if success_count > 0:
            # Only mark successful values
            with stats.lock:
                for value in success_values:
                    if field == 'name':
                        stats.successful_name_tags.add(value)
                    else:
                        stats.successful_ip_tags.add(value)
                
                for value in failed_values:
                    if field == 'name':
                        stats.failed_name_tags.add(value)
                    else:
                        stats.failed_ip_tags.add(value)
            
            success_ratio = success_count / len(all_device_ids) if all_device_ids else 0
            return success_ratio >= 0.5  # Consider batch successful if at least 50% were tagged
        
        # If tag assignment fails, mark all cells as failed
        for cell in cells:
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
        return False
    finally:
        return_session(session)

class BatchSizeManager:
    def __init__(self, initial_size=BATCH_SIZE_INITIAL, min_size=10, max_size=MAX_BATCH_SIZE):
        self.current_size = initial_size
        self.min_size = min_size
        self.max_size = max_size
        self.success_count = 0
        self.failure_count = 0
        self.lock = Lock()

    def adjust_batch_size(self, success):
        with self.lock:
            if success:
                self.success_count += 1
                self.failure_count = 0
                if self.success_count >= 3:
                    self.increase_batch_size()
            else:
                self.failure_count += 1
                self.success_count = 0
                if self.failure_count >= 2:
                    self.decrease_batch_size()

    def increase_batch_size(self):
        new_size = min(self.current_size * 1.5, self.max_size)
        if new_size != self.current_size:
            self.current_size = int(new_size)
            logging.info(f"Increased batch size to {self.current_size}")

    def decrease_batch_size(self):
        new_size = max(self.current_size * 0.5, self.min_size)
        if new_size != self.current_size:
            self.current_size = int(new_size)
            logging.info(f"Decreased batch size to {self.current_size}")

    def get_size(self):
        with self.lock:
            return self.current_size

def preload_tag_cache(sheet, tag_col):
    """Pre-load all unique tags from the sheet to optimize tag creation."""
    unique_tags = set()
    for row in sheet.iter_rows(min_row=2):
        tag_name = row[tag_col - 1].value
        if tag_name:
            unique_tags.add(tag_name)
    
    logging.info(f"Found {len(unique_tags)} unique tags to process")
    return unique_tags

def process_excel_with_pandas():
    """Use pandas to read Excel file for better memory efficiency."""
    logging.info(f"Reading Excel file {EXCEL_FILE} with pandas")
    
    # Read Excel using pandas for better memory efficiency
    df = pd.read_excel(EXCEL_FILE)
    
    # Create a new Excel workbook for updated data
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx+2, column=col_idx, value=value)
    
    # Get column indices
    name_col, ipaddr_col, tag_col = None, None, None
    for idx, col_name in enumerate(df.columns):
        if col_name.lower() == 'name':
            name_col = idx + 1
        elif col_name.lower() == 'ipaddr':
            ipaddr_col = idx + 1
        elif col_name.lower() == 'tag':
            tag_col = idx + 1
    
    return wb, ws, name_col, ipaddr_col, tag_col

def process_excel_sheet(sheet, headers):
    """Process each row for both name and ipaddr columns with optimized batch processing."""
    name_col, ipaddr_col, tag_col = None, None, None
    for cell in sheet[1]:
        if cell.value and cell.value.lower() == 'name':
            name_col = cell.column
        elif cell.value and cell.value.lower() == 'ipaddr':
            ipaddr_col = cell.column
        elif cell.value and cell.value.lower() == 'tag':
            tag_col = cell.column
    
    # Validate required columns
    if not tag_col:
        logging.error("Tag column not found in Excel sheet. Aborting.")
        return
    
    if not name_col and not ipaddr_col:
        logging.error("Neither name nor ipaddr column found in Excel sheet. Aborting.")
        return
    
    # Preload tags and cache them
    preload_tags = preload_tag_cache(sheet, tag_col)
    for tag_name in preload_tags:
        tag_id = get_tag_id(tag_name, headers)
        if not tag_id:
            tag_id = create_tag(tag_name, headers)
            if not tag_id:
                logging.error(f"Failed to create tag: {tag_name}")
    
    logging.info(f"Preloaded {len(stats.tag_cache)} tags into cache")

    # Group rows by tag to reduce API calls
    tag_groups = defaultdict(list)
    
    previous_tag = None
    max_row = sheet.max_row
    
    # First pass: group rows by tag
    logging.info(f"Grouping {max_row-1} rows by tag")
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
        if row_idx % 1000 == 0:
            logging.info(f"Processed {row_idx}/{max_row} rows for grouping")
        
        tag_name = row[tag_col - 1].value
        
        if not tag_name and previous_tag:
            tag_name = previous_tag
            row[tag_col - 1].value = tag_name
        
        if tag_name:
            previous_tag = tag_name
            tag_groups[tag_name].append(row)
    
    logging.info(f"Grouped rows into {len(tag_groups)} unique tags")
    
    # Process tags in parallel
    semaphore = Semaphore(MAX_WORKERS)
    batch_manager = BatchSizeManager()
    
    def process_batch(values, field, cells, tag_id):
        with semaphore:
            success = batch_search_and_tag_devices(values, field, tag_id, headers, cells)
            batch_manager.adjust_batch_size(success)
            return success

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = []
        
        # Process each tag group
        for tag_name, rows in tag_groups.items():
            logging.info(f"Processing tag: {tag_name} with {len(rows)} rows")
            
            tag_id = stats.tag_cache.get(tag_name)
            if not tag_id:
                tag_id = get_tag_id(tag_name, headers) or create_tag(tag_name, headers)
                if not tag_id:
                    logging.error(f"Skipping group, unable to retrieve or create tag: {tag_name}")
                    continue
            
            # Group by field (name or IP)
            name_values = []
            name_cells = []
            ip_values = []
            ip_cells = []
            
            for row in rows:
                # Process name column
                if name_col is not None and row[name_col - 1].value:
                    name_values.append(row[name_col - 1].value)
                    name_cells.append(row[name_col - 1])
                
                # Process ipaddr column
                if ipaddr_col is not None and row[ipaddr_col - 1].value:
                    ip_values.append(row[ipaddr_col - 1].value)
                    ip_cells.append(row[ipaddr_col - 1])
            
            # Process names in batches
            for i in range(0, len(name_values), batch_manager.get_size()):
                batch_values = name_values[i:i + batch_manager.get_size()]
                batch_cells = name_cells[i:i + batch_manager.get_size()]
                futures.append(executor.submit(process_batch, batch_values, 'name', batch_cells, tag_id))
            
            # Process IPs in batches
            for i in range(0, len(ip_values), batch_manager.get_size()):
                batch_values = ip_values[i:i + batch_manager.get_size()]
                batch_cells = ip_cells[i:i + batch_manager.get_size()]
                futures.append(executor.submit(process_batch, batch_values, 'ipaddr', batch_cells, tag_id))
        
        # Monitor progress
        total_futures = len(futures)
        logging.info(f"Submitted {total_futures} batch tasks for processing")
        
        # Wait for tasks to complete with progress tracking
        for i, future in enumerate(concurrent.futures.as_completed(futures)):
            if (i + 1) % 10 == 0 or (i + 1) == total_futures:
                logging.info(f"Completed {i + 1}/{total_futures} batch tasks")

def print_summary():
    """Print a summary of all tagging operations."""
    duration = datetime.now() - stats.start_time
    minutes, seconds = divmod(duration.total_seconds(), 60)
    hours, minutes = divmod(minutes, 60)
    
    logging.info("\n=== Tagging Operation Summary ===")
    logging.info(f"Total processing time: {int(hours)}h {int(minutes)}m {int(seconds)}s")
    logging.info(f"Total processed items: {stats.processed_count}")
    if duration.total_seconds() > 0:
        rate = stats.processed_count / duration.total_seconds()
        logging.info(f"Processing rate: {rate:.2f} items/second")
    
    logging.info(f"Created new tags: {len(stats.created_tags)}")
    logging.info(f"Used existing tags: {len(stats.existing_tags)}")
    logging.info(f"Successfully tagged by name: {len(stats.successful_name_tags)}")
    logging.info(f"Failed to tag by name: {len(stats.failed_name_tags)}")
    logging.info(f"Successfully tagged by IP: {len(stats.successful_ip_tags)}")
    logging.info(f"Failed to tag by IP: {len(stats.failed_ip_tags)}")
    
    if stats.failed_name_tags and len(stats.failed_name_tags) <= 100:
        logging.info("\nFailed name tags (first 100):")
        for name in list(stats.failed_name_tags)[:100]:
            logging.info(f"- {name}")
    elif stats.failed_name_tags:
        logging.info(f"\nFailed name tags: {len(stats.failed_name_tags)} items")
    
    if stats.failed_ip_tags and len(stats.failed_ip_tags) <= 100:
        logging.info("\nFailed IP tags (first 100):")
        for ip in list(stats.failed_ip_tags)[:100]:
            logging.info(f"- {ip}")
    elif stats.failed_ip_tags:
        logging.info(f"\nFailed IP tags: {len(stats.failed_ip_tags)} items")

def main():
    log_filename = setup_logging()
    logging.info("Starting tagging operations...")
    logging.info(f"Processing file: {EXCEL_FILE} with {MAX_WORKERS} workers")
    stats.start_time = datetime.now()
    
    # Initialize session pool
    logging.info(f"Initializing session pool with {SESSION_POOL_SIZE} connections")
    for _ in range(SESSION_POOL_SIZE):
        session_pool.append(create_session())
    
    headers = {"Authorization": get_auth_header()}
    if not headers["Authorization"]:
        logging.error("Authentication failed. Exiting...")
        return
    
    # Check file size and use appropriate method
    file_size_mb = os.path.getsize(EXCEL_FILE) / (1024 * 1024)
    logging.info(f"Excel file size: {file_size_mb:.2f} MB")
    
    if file_size_mb > 100:  # For very large files
        try:
            wb, sheet, name_col, ipaddr_col, tag_col = process_excel_with_pandas()
            process_excel_sheet(sheet, headers)
            wb.save("updated_" + EXCEL_FILE)
        except Exception as e:
            logging.error(f"Error processing large file with pandas: {str(e)}")
            logging.info("Falling back to regular openpyxl processing")
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet = wb.active
            process_excel_sheet(sheet, headers)
            wb.save("updated_" + EXCEL_FILE)
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        process_excel_sheet(sheet, headers)
        wb.save("updated_" + EXCEL_FILE)
    
    print_summary()
    logging.info(f"Workbook saved as 'updated_{EXCEL_FILE}'")
    logging.info(f"Detailed log saved as '{log_filename}'")

if __name__ == "__main__":
    main()
