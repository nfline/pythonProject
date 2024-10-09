import socket
import openpyxl

# 读取Excel文件中的IP地址
def read_ips_from_excel(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    ips = []

    # 假设IP地址在第一列，从第二行开始读取（跳过表头）
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        ip = row[0]
        if ip:
            ips.append(ip)
    return ips

# 获取IP地址对应的主机名
def get_hostname(ip):
    try:
        hostname = socket.gethostbyaddr(ip)[0]
    except socket.herror:
        hostname = 'Hostname not found'
    return hostname

# 将IP和主机名写入新的Excel文件
def write_to_excel(output_file, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "IP to Hostname"

    # 写表头
    sheet.append(["IP Address", "Hostname"])

    # 写数据
    for ip, hostname in data:
        sheet.append([ip, hostname])

    wb.save(output_file)

# 主函数
def main(input_file, sheet_name, output_file):
    ips = read_ips_from_excel(input_file, sheet_name)
    data = []

    for ip in ips:
        hostname = get_hostname(ip)
        data.append((ip, hostname))

    write_to_excel(output_file, data)
    print(f"Results written to {output_file}")

# 使用示例
if __name__ == "__main__":
    input_file = "input_ips.xlsx"  # 输入Excel文件
    sheet_name = "Sheet1"          # 工作表名称
    output_file = "output_hostnames.xlsx"  # 输出Excel文件

    main(input_file, sheet_name, output_file)
